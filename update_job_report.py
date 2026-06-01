#!/usr/bin/env python3
"""
update_job_report.py
--------------------
Scans a folder of resume / application files and builds an editable job-application
tracker (job_applications.csv).

What it does
============
* Pulls Company, Job Title and Date from each application file's name + date.
* Writes ONE row per job (a resume's .pdf and .docx collapse into a single row).
* Skips cover letters and interview-prep files so they don't create duplicate rows
  (a "Cover Letter? = Yes" flag is set instead when a matching cover letter exists).
* Is safe to re-run on future dates: it reads the existing job_applications.csv,
  KEEPS everything you've edited (Date Applied, Company, Job Title, Status, Notes),
  adds any newly-found jobs, and refreshes the auto columns.
* Backs up the previous report into report_backups/ before each save.

How to run
==========
    python update_job_report.py
        -> uses the folder this script lives in

    python update_job_report.py "C:\\path\\to\\folder"
        -> uses a folder you point it at

Which columns you can edit
==========================
    Date Applied, Company, Job Title, Status, Notes   <- edit freely; kept on re-run
    Cover Letter?, Source Files, First Added, Last Updated   <- auto-managed
(Edits are matched back to files via the "Source Files" column, so changing the
 Company/Title text in the sheet will NOT cause a duplicate next run.)
"""

import csv
import os
import re
import shutil
import sys
from datetime import datetime

# ---------------------------------------------------------------- configuration
REPORT_NAME   = "job_applications.csv"
XLSX_NAME     = "job_applications.xlsx"
BACKUP_DIR    = "report_backups"
APP_EXTS      = {".pdf", ".docx", ".doc"}
DEFAULT_STATUS = "Applied"

# Options offered in the Status dropdown of the .xlsx, each with (fill, font) colors.
# Order here is also the order shown in the summary legend.
STATUS_STYLES = {
    "To Apply":     ("E7E6E6", "7F7F7F"),
    "Applied":      ("DDEBF7", "2E75B6"),
    "Phone Screen": ("D6EAF8", "1F6391"),
    "Interviewing": ("FFF2CC", "BF8F00"),
    "Final Round":  ("FCE4D6", "C55A11"),
    "Offer":        ("E2EFDA", "548235"),
    "Accepted":     ("C6EFCE", "006100"),
    "Rejected":     ("FFC7CE", "9C0006"),
    "Withdrawn":    ("F2F2F2", "808080"),
    "Ghosted":      ("EAE3F2", "7030A0"),
}
STATUS_OPTIONS = list(STATUS_STYLES.keys())

# Files we never turn into their own job row
COVER_PATTERN  = re.compile(r"cover", re.IGNORECASE)      # cover letters
SKIP_PATTERNS  = (
    COVER_PATTERN,
    re.compile(r"interview", re.IGNORECASE),              # interview prep
)

# ============================================================================
#  >>> STEP 1 (REQUIRED): put YOUR full name here, exactly as it appears at the
#      START of your resume file names. First + last is enough; extra middle
#      names are fine. This is how the script knows where your name ends and the
#      company name begins.  Example:  YOUR_NAME = "Jane Doe"
# ============================================================================
YOUR_NAME = "First Last"

_name_tokens = [re.escape(t) for t in YOUR_NAME.split()]
NAME_PREFIX = re.compile(r"^" + r"[ _]*".join(_name_tokens) + r"[ _\-]*", re.IGNORECASE)
ACRONYMS    = {"AI", "TPM", "SHI", "HR", "IT", "US", "ML"}


# ---------------------------------------------------------------- name parsing
def split_camel(s: str) -> str:
    """Insert spaces into CamelCase: 'AIEnablementLead' -> 'AI Enablement Lead'."""
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", s)
    s = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", s)
    return s


def prettify(s: str) -> str:
    """Turn a filename chunk into a human-readable Company or Title."""
    s = s.replace("_", " ").replace("-", " ")
    s = split_camel(s)
    out = []
    for w in s.split():
        if w.upper() in ACRONYMS:
            out.append(w.upper())
        elif w.isupper() and len(w) > 1:   # keep brand caps like DISCO, AHEAD, SHI
            out.append(w)
        else:
            out.append(w[:1].upper() + w[1:])
    res = " ".join(out)
    res = re.sub(r"\bSr\b", "Senior", res)
    res = re.sub(r"\bAi\b", "AI", res)
    return res.strip()


def parse_company_title(stem: str):
    """Strip the name prefix, then split the remainder into (company, title)."""
    rest = NAME_PREFIX.sub("", stem).strip(" _-")
    # split on the first '_' or '-' that separates company from title
    m = re.search(r"[_\-]", rest)
    if m:
        company = rest[: m.start()]
        title = rest[m.end():]
    else:
        company = ""          # company not encoded in the filename
        title = rest
    return prettify(company), prettify(title)


def is_skipped(name: str) -> bool:
    return any(p.search(name) for p in SKIP_PATTERNS)


def group_key(company: str, title: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (company + title).lower())


# ---------------------------------------------------------------- scanning
def scan_folder(folder: str):
    """Return {group_key: job_dict} for every application (non-cover-letter) file."""
    jobs = {}
    cover_companies = set()   # lowercased company names that have a cover letter

    for entry in os.scandir(folder):
        if not entry.is_file():
            continue
        name = entry.name
        stem, ext = os.path.splitext(name)
        if ext.lower() not in APP_EXTS:
            continue

        company, title = parse_company_title(stem)

        if is_skipped(name):
            if company and COVER_PATTERN.search(name):   # only real cover letters
                cover_companies.add(company.lower())
            continue

        key = group_key(company, title)
        if not key:
            continue
        mtime = datetime.fromtimestamp(entry.stat().st_mtime)

        job = jobs.get(key)
        if job is None:
            jobs[key] = {
                "company": company,
                "title": title,
                "date": mtime,
                "files": {name},
            }
        else:
            job["files"].add(name)
            if mtime < job["date"]:        # earliest file = when the app was created
                job["date"] = mtime

    # flag jobs whose company also has a cover letter on disk
    for job in jobs.values():
        job["cover"] = "Yes" if job["company"].lower() in cover_companies else "No"
    return jobs


# ---------------------------------------------------------------- merge + write
FIELDS = [
    "Date Applied", "Company", "Job Title", "Status", "Notes",
    "Cover Letter?", "Source Files", "First Added", "Last Updated",
]


def load_existing(folder: str):
    """Read prior edits, preferring the .xlsx (what you edit) then the .csv."""
    xlsx_path = os.path.join(folder, XLSX_NAME)
    csv_path = os.path.join(folder, REPORT_NAME)
    if os.path.exists(xlsx_path):
        rows = load_existing_xlsx(xlsx_path)
        if rows:
            return rows
    if os.path.exists(csv_path):
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    return []


def load_existing_xlsx(path: str):
    """Read data rows out of the formatted workbook (skips the summary block)."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return []
    ws = wb.active
    header_row = None
    headers = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in r]
        if "Date Applied" in vals and "Company" in vals:
            header_row = r[0].row
            headers = [str(v).strip() if v is not None else "" for v in vals]
            break
    if header_row is None:
        return []
    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in r):
            continue
        row = {}
        for h, v in zip(headers, r):
            if not h:
                continue
            row[h] = "" if v is None else str(v).strip()
        if row.get("Date Applied") or row.get("Company") or row.get("Source Files"):
            rows.append(row)
    return rows


def files_of(row: dict):
    return {x.strip() for x in (row.get("Source Files") or "").split(";") if x.strip()}


def merge(existing_rows, jobs, today):
    """Combine scanned jobs with the previous report, preserving user edits."""
    used = [False] * len(existing_rows)
    result = []

    for key, job in jobs.items():
        new_files = job["files"]
        match = None
        for i, row in enumerate(existing_rows):
            if not used[i] and files_of(row) & new_files:   # share a file -> same job
                match = i
                break

        if match is not None:
            row = dict(existing_rows[match])
            used[match] = True
            all_files = files_of(row) | new_files
            changed = all_files != files_of(row)
            # auto columns refresh; user columns kept as-is
            row["Source Files"] = "; ".join(sorted(all_files))
            row["Cover Letter?"] = job["cover"]
            if not row.get("First Added"):
                row["First Added"] = today
            if changed or not row.get("Last Updated"):
                row["Last Updated"] = today
            # backfill anything blank
            row.setdefault("Status", DEFAULT_STATUS)
            if not row.get("Date Applied"):
                row["Date Applied"] = job["date"].strftime("%Y-%m-%d")
            if not row.get("Company"):
                row["Company"] = job["company"]
            if not row.get("Job Title"):
                row["Job Title"] = job["title"]
        else:
            row = {
                "Date Applied": job["date"].strftime("%Y-%m-%d"),
                "Company": job["company"],
                "Job Title": job["title"],
                "Status": DEFAULT_STATUS,
                "Notes": "",
                "Cover Letter?": job["cover"],
                "Source Files": "; ".join(sorted(new_files)),
                "First Added": today,
                "Last Updated": today,
            }
        result.append(row)

    # keep old rows whose files are gone (e.g. you moved a file) so notes survive
    for i, row in enumerate(existing_rows):
        if not used[i]:
            r = {k: row.get(k, "") for k in FIELDS}
            note = r.get("Notes", "")
            flag = "[file not found in folder]"
            if flag not in note:
                r["Notes"] = (note + "  " + flag).strip()
            result.append(r)

    # newest applications first; blank dates sink to the bottom
    result.sort(key=lambda r: r.get("Date Applied") or "0000-00-00", reverse=True)
    return result


def write_report(path, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in FIELDS})


def status_summary(rows):
    """Return (total, [(status, count), ...]) for the summary block."""
    counts = {}
    for r in rows:
        s = (r.get("Status") or "").strip() or "(blank)"
        counts[s] = counts.get(s, 0) + 1
    order = {s: i for i, s in enumerate(STATUS_OPTIONS)}
    ordered = sorted(counts.items(), key=lambda kv: (order.get(kv[0], 99), kv[0]))
    return len(rows), ordered


def _resume_link(row):
    """Pick the best file to link a row to (prefer the PDF resume)."""
    files = [f.strip() for f in (row.get("Source Files") or "").split(";") if f.strip()]
    pdfs = [f for f in files if f.lower().endswith(".pdf")]
    return (pdfs or files or [None])[0]


def write_xlsx(path, rows, today):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.sheet_view.showGridLines = False

    navy   = "1F3864"
    blue   = "2E5496"
    stripe = "F2F6FC"
    link   = "0563C1"
    ncols  = len(FIELDS)
    last_col = get_column_letter(ncols)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, size=20, color="FFFFFF")
    sub_font    = Font(italic=True, size=10, color="D9E1F2")
    bold        = Font(bold=True, color=navy)
    wrap        = Alignment(wrap_text=True, vertical="top")
    topleft     = Alignment(vertical="top")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="D6DCE4")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    total, summary = status_summary(rows)

    # ---- title banner (rows 1-2, merged across all columns) ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    t = ws.cell(row=1, column=1, value="  Job Application Tracker")
    t.font = title_font
    t.alignment = Alignment(vertical="center")
    s = ws.cell(row=2, column=1,
                value=f"  {total} applications  ·  last updated {today}")
    s.font = sub_font
    for rr in (1, 2):
        for col in range(1, ncols + 1):
            ws.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=navy)
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 18

    # ---- status legend (row 4): colored pills with counts ----
    legend_row = 4
    lab = ws.cell(row=legend_row, column=1, value="Status:")
    lab.font = bold
    col = 2
    count_map = dict(summary)
    for status in STATUS_OPTIONS:
        n = count_map.get(status, 0)
        if n == 0:
            continue
        fill_hex, font_hex = STATUS_STYLES[status]
        c = ws.cell(row=legend_row, column=col, value=f"{status}  {n}")
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.font = Font(bold=True, color=font_hex)
        c.alignment = center
        c.border = Border(left=Side(style="thin", color=fill_hex),
                          right=Side(style="thin", color=fill_hex),
                          top=Side(style="thin", color=fill_hex),
                          bottom=Side(style="thin", color=fill_hex))
        col += 1
    ws.row_dimensions[legend_row].height = 20

    header_row = legend_row + 2   # spacer row, then the table header

    # ---- table header ----
    for col, name in enumerate(FIELDS, start=1):
        c = ws.cell(row=header_row, column=col, value=name)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor=blue)
        c.alignment = Alignment(vertical="center",
                                horizontal="center" if name == "Cover Letter?" else "left")
        c.border = border
    ws.row_dimensions[header_row].height = 22

    # ---- data rows ----
    for i, r in enumerate(rows):
        excel_row = header_row + 1 + i
        for col, name in enumerate(FIELDS, start=1):
            c = ws.cell(row=excel_row, column=col, value=r.get(name, ""))
            c.border = border
            c.alignment = wrap if name in ("Notes", "Source Files") else topleft
            if name == "Cover Letter?":
                c.alignment = center
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=stripe)
        # make the Company cell a clickable link to the resume PDF
        target = _resume_link(r)
        if target:
            cc = ws.cell(row=excel_row, column=FIELDS.index("Company") + 1)
            cc.hyperlink = target
            cc.font = Font(color=link, underline="single", bold=True)

    first_data = header_row + 1
    last_data = header_row + len(rows)
    status_col = get_column_letter(FIELDS.index("Status") + 1)

    if len(rows) > 0:
        rng = f"{status_col}{first_data}:{status_col}{last_data}"
        # ---- Status dropdown ----
        dv = DataValidation(type="list",
                            formula1='"' + ",".join(STATUS_OPTIONS) + '"',
                            allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(rng)
        # ---- live color-coding: recolors instantly when you pick a status ----
        for status, (fill_hex, font_hex) in STATUS_STYLES.items():
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=[f'"{status}"'],
                fill=PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid"),
                font=Font(bold=True, color=font_hex)))

    # ---- widths, freeze, filter ----
    widths = {
        "Date Applied": 13, "Company": 20, "Job Title": 32, "Status": 14,
        "Notes": 42, "Cover Letter?": 12, "Source Files": 52,
        "First Added": 12, "Last Updated": 12,
    }
    for col, name in enumerate(FIELDS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(name, 16)

    ws.freeze_panes = ws.cell(row=first_data, column=1)
    if len(rows) > 0:
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data}"

    wb.save(path)


# ---------------------------------------------------------------- main
def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else os.path.dirname(os.path.abspath(__file__))
    folder = os.path.abspath(folder)
    if not os.path.isdir(folder):
        sys.exit(f"Folder not found: {folder}")

    report_path = os.path.join(folder, REPORT_NAME)
    xlsx_path = os.path.join(folder, XLSX_NAME)
    today = datetime.now().strftime("%Y-%m-%d")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # back up the previous versions before touching them
    bdir = os.path.join(folder, BACKUP_DIR)
    for src, suffix in ((report_path, "csv"), (xlsx_path, "xlsx")):
        if os.path.exists(src):
            os.makedirs(bdir, exist_ok=True)
            shutil.copy2(src, os.path.join(bdir, f"job_applications_{stamp}.{suffix}"))

    existing = load_existing(folder)
    jobs = scan_folder(folder)
    rows = merge(existing, jobs, today)
    write_report(report_path, rows)
    try:
        write_xlsx(xlsx_path, rows, today)
        xlsx_ok = True
    except Exception as e:                      # never let xlsx issues lose the csv
        xlsx_ok = False
        print(f"WARNING: could not write .xlsx ({e}); .csv is still updated.")

    print(f"Folder scanned : {folder}")
    print(f"Jobs found     : {len(jobs)}")
    print(f"Rows in report : {len(rows)} (kept {len(existing)} existing)")
    print(f"CSV  written   : {report_path}")
    if xlsx_ok:
        print(f"XLSX written   : {xlsx_path}")
    if os.path.exists(bdir):
        print(f"Backup saved   : {bdir}")


if __name__ == "__main__":
    main()
